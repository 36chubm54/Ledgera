from __future__ import annotations

import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from domain.debt import Debt
from domain.reports import Report
from services.analytics.report import build_tag_group_reports
from utils.export.i18n import (
    balance_label as localized_balance_label,
)
from utils.export.i18n import (
    category_breakdown_headers,
    category_breakdown_unavailable,
    category_title,
    debt_headers,
    debt_kind_label,
    debt_status_label,
    final_balance_label,
    fixed_amounts_note,
    grouped_category_totals_note,
    grouped_report_csv_headers,
    monthly_summary_headers,
    report_xlsx_headers,
    sheet_title_by_category,
    sheet_title_by_tag,
    sheet_title_debts,
    sheet_title_report,
    sheet_title_warnings,
    sheet_title_yearly,
    subtotal_label,
    tag_title,
    total_label,
    warnings_header,
)
from utils.export.i18n import (
    statement_title as localized_statement_title,
)
from utils.export.io import save_workbook_output
from utils.finance.debt_report import debt_progress_percent, debts_for_report_period
from utils.spreadsheets.excel_export_builders import (
    build_category_sheet_sections,
    build_monthly_summary_rows,
    build_report_sheet_rows,
    build_tag_sheet_sections,
)
from utils.spreadsheets.excel_sheet_styles import (
    SECTION_FILL,
    SUBTOTAL_FILL,
    THIN_BORDER,
    TOTAL_FILL,
)
from utils.spreadsheets.excel_sheet_styles import (
    set_auto_width as _set_auto_width,
)
from utils.spreadsheets.excel_sheet_styles import (
    style_data_row as _style_data_row,
)
from utils.spreadsheets.excel_sheet_styles import (
    style_header_row as _style_header_row,
)
from utils.spreadsheets.excel_sheet_styles import (
    style_title_row as _style_title_row,
)
from utils.spreadsheets.excel_sheet_styles import (
    style_total_row as _style_total_row,
)

logger = logging.getLogger(__name__)


def grouped_sections_with_warning(report: Report) -> tuple[dict[str, Report], str | None]:
    try:
        return report.grouped_by_category(), None
    except (AttributeError, TypeError, ValueError, RuntimeError) as exc:
        logger.warning("Failed to build grouped report sections for XLSX export: %s", exc)
        return {}, category_breakdown_unavailable(exc)


def _append_warning_sheet(wb: Workbook, message: str) -> None:
    ws = wb.create_sheet(title=sheet_title_warnings())
    ws.append([warnings_header()])
    _style_header_row(ws, 1)
    ws.append([message])
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = min(max(len(message) // 2, 24), 120)


def _should_add_by_category_sheet(report: Report, groups: dict[str, Report]) -> bool:
    if len(groups) > 1:
        return True
    if len(groups) != 1:
        return False
    only_subreport = next(iter(groups.values()))
    return len(list(only_subreport.records())) < len(list(report.records()))


def _should_add_by_tag_sheet(report: Report, groups: dict[str, Report]) -> bool:
    if len(groups) > 1:
        return True
    if len(groups) != 1:
        return False
    only_subreport = next(iter(groups.values()))
    return len(list(only_subreport.records())) < len(list(report.display_records()))


def _append_debts_sheet(wb: Workbook, debts: list[Debt]) -> None:
    if not debts:
        return
    ws = wb.create_sheet(sheet_title_debts(), index=len(wb.sheetnames))
    ws.append(debt_headers())
    _style_header_row(ws, 1)
    ws.freeze_panes = "A2"
    for debt in debts:
        settled = (int(debt.total_amount_minor) - int(debt.remaining_amount_minor)) / 100.0
        ws.append(
            [
                str(debt.contact_name),
                debt_kind_label(str(debt.kind.value)),
                debt_status_label(str(debt.status.value)),
                str(debt.created_at),
                str(debt.closed_at or "-"),
                str(debt.currency).upper(),
                debt.total_amount_minor / 100.0,
                debt.remaining_amount_minor / 100.0,
                settled,
                debt_progress_percent(debt),
            ]
        )
        _style_data_row(ws, ws.max_row, amount_columns=(7, 8, 9, 10))
        ws.cell(row=ws.max_row, column=10).number_format = "0.00"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"
    _set_auto_width(ws)


def report_to_xlsx(
    report: Report,
    filepath: str,
    *,
    debts: list[Debt] | None = None,
    base_currency: str = "KZT",
) -> None:
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = sheet_title_report()
        ws.append([localized_statement_title(report.statement_title), "", "", "", ""])
        _style_title_row(ws, 1, columns=5)
        ws.append(report_xlsx_headers(base_currency))
        _style_header_row(ws, 2)
        ws.append(["", "", "", "", fixed_amounts_note()])
        ws["D3"].alignment = Alignment(horizontal="right", vertical="center")
        ws["D3"].font = Font(italic=True, color="666666")
        ws.freeze_panes = "A3"

    if ws is not None:
        report_rows = build_report_sheet_rows(report)
        for index, row in enumerate(report_rows):
            if index == 0 and row[1] == report.balance_label:
                styled_row = ["", localized_balance_label(report.balance_label), "", row[3], ""]
                ws.append(styled_row)
                _style_total_row(ws, ws.max_row, fill=SECTION_FILL, amount_columns=(4,))
                continue
            ws.append(row)
            if row[0] == subtotal_label():
                _style_total_row(ws, ws.max_row, fill=SUBTOTAL_FILL, amount_columns=(4,))
            elif row[0] == final_balance_label():
                _style_total_row(ws, ws.max_row, fill=TOTAL_FILL, amount_columns=(4,))
            else:
                _style_data_row(ws, ws.max_row, amount_columns=(4,))
        ws.auto_filter.ref = f"A2:E{ws.max_row}"
        _set_auto_width(ws)

    summary_year, monthly_rows = report.monthly_income_expense_rows()
    summary_ws = wb.create_sheet(sheet_title_yearly())
    if summary_ws is not None:
        summary_ws.append(monthly_summary_headers(summary_year, base_currency))
        _style_header_row(summary_ws, 1, center=True)
        summary_ws.freeze_panes = "A2"
        summary_rows, _total_income, _total_expense = build_monthly_summary_rows(monthly_rows)
        for row in summary_rows:
            summary_ws.append(row)
            _style_data_row(summary_ws, summary_ws.max_row, amount_columns=(2, 3))
        _style_total_row(summary_ws, summary_ws.max_row, fill=TOTAL_FILL, amount_columns=(2, 3))
        summary_ws.auto_filter.ref = f"A1:C{summary_ws.max_row}"
        _set_auto_width(summary_ws)

    groups, grouped_warning = grouped_sections_with_warning(report)
    if _should_add_by_category_sheet(report, groups):
        bycat_ws = wb.create_sheet(title=sheet_title_by_category(), index=1)
        for category, section_rows, _records_total in build_category_sheet_sections(groups):
            bycat_ws.append([category_title(category)])
            category_row = bycat_ws.max_row
            bycat_ws.merge_cells(
                start_row=category_row,
                start_column=1,
                end_row=category_row,
                end_column=3,
            )
            bycat_ws.cell(row=category_row, column=1).font = Font(bold=True, color="1F1F1F")
            bycat_ws.cell(row=category_row, column=1).fill = SECTION_FILL
            bycat_ws.cell(row=category_row, column=1).border = THIN_BORDER
            bycat_ws.append(category_breakdown_headers(base_currency))
            _style_header_row(bycat_ws, bycat_ws.max_row)
            for row in section_rows:
                bycat_ws.append(row)
                if row[0] == subtotal_label():
                    _style_total_row(
                        bycat_ws,
                        bycat_ws.max_row,
                        fill=SUBTOTAL_FILL,
                        amount_columns=(3,),
                    )
                else:
                    _style_data_row(bycat_ws, bycat_ws.max_row, amount_columns=(3,))
            bycat_ws.append([""])
        bycat_ws.freeze_panes = "A2"
        _set_auto_width(bycat_ws)
    elif grouped_warning:
        _append_warning_sheet(wb, grouped_warning)

    tag_groups = build_tag_group_reports(report)
    if _should_add_by_tag_sheet(report, tag_groups):
        bytag_ws = wb.create_sheet(title=sheet_title_by_tag(), index=2)
        for tag, section_rows, _records_total in build_tag_sheet_sections(tag_groups):
            bytag_ws.append([tag_title(tag)])
            tag_row = bytag_ws.max_row
            bytag_ws.merge_cells(
                start_row=tag_row,
                start_column=1,
                end_row=tag_row,
                end_column=5,
            )
            bytag_ws.cell(row=tag_row, column=1).font = Font(bold=True, color="1F1F1F")
            bytag_ws.cell(row=tag_row, column=1).fill = SECTION_FILL
            bytag_ws.cell(row=tag_row, column=1).border = THIN_BORDER
            bytag_ws.append(report_xlsx_headers(base_currency))
            _style_header_row(bytag_ws, bytag_ws.max_row)
            for row in section_rows:
                bytag_ws.append(row)
                if row[0] == subtotal_label():
                    _style_total_row(
                        bytag_ws,
                        bytag_ws.max_row,
                        fill=SUBTOTAL_FILL,
                        amount_columns=(4,),
                    )
                else:
                    _style_data_row(bytag_ws, bytag_ws.max_row, amount_columns=(4,))
            bytag_ws.append([""])
        bytag_ws.freeze_panes = "A2"
        _set_auto_width(bytag_ws)

    _append_debts_sheet(wb, debts_for_report_period(report, list(debts or [])))
    save_workbook_output(wb, filepath, context="report XLSX export", logger=logger)


def grouped_report_to_xlsx(
    statement_title: str,
    grouped_rows: list[tuple[str, int, float]],
    filepath: str,
    *,
    base_currency: str = "KZT",
) -> None:
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = sheet_title_report()
        ws.append([statement_title, "", ""])
        _style_title_row(ws, 1, columns=3)
        ws.append(grouped_report_csv_headers(base_currency))
        _style_header_row(ws, 2)
        ws.append(["", "", grouped_category_totals_note()])
        ws["C3"].alignment = Alignment(horizontal="right", vertical="center")
        ws["C3"].font = Font(italic=True, color="666666")
        ws.freeze_panes = "A3"

        total_base = 0.0
        for category, operations_count, amount_base in grouped_rows:
            total_base += float(amount_base)
            ws.append([category, int(operations_count), float(amount_base)])
            _style_data_row(ws, ws.max_row, amount_columns=(3,))

        ws.append([total_label(), "", total_base])
        _style_total_row(ws, ws.max_row, fill=TOTAL_FILL, amount_columns=(3,))
        ws.auto_filter.ref = f"A2:C{ws.max_row}"
        _set_auto_width(ws)

    save_workbook_output(wb, filepath, context="grouped report XLSX export", logger=logger)
