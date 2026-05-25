from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SUBTOTAL_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def style_title_row(ws, row_idx: int, *, columns: int) -> None:
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=columns)
    cell = ws.cell(row=row_idx, column=1)
    cell.font = Font(bold=True, size=14, color="1F1F1F")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header_row(ws, row_idx: int, *, center: bool = False) -> None:
    for cell in ws[row_idx]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center",
        )


def style_data_row(ws, row_idx: int, *, amount_columns: tuple[int, ...] = ()) -> None:
    for cell in ws[row_idx]:
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for column_idx in amount_columns:
        ws.cell(row=row_idx, column=column_idx).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=column_idx).alignment = Alignment(
            horizontal="right",
            vertical="center",
        )


def style_total_row(
    ws, row_idx: int, *, fill: PatternFill, amount_columns: tuple[int, ...]
) -> None:
    for cell in ws[row_idx]:
        cell.font = Font(bold=True, color="1F1F1F")
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for column_idx in amount_columns:
        ws.cell(row=row_idx, column=column_idx).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=column_idx).alignment = Alignment(
            horizontal="right",
            vertical="center",
        )


def set_auto_width(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), len(value))
    for column_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(column_idx)].width = min(max(width + 2, 12), 32)
