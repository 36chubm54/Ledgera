from __future__ import annotations

import os
from collections.abc import Callable
from logging import Logger
from typing import Any

from openpyxl import load_workbook

from domain.import_policy import ImportPolicy
from domain.records import MandatoryExpenseRecord, Record
from utils.export.i18n import (
    canonical_report_header,
    canonical_report_row_type,
    is_report_total_row_label,
    is_statement_title,
)
from utils.finance.money import to_money_float
from utils.import_core import ImportSummary, norm_key, safe_type
from utils.records.tabular import resolve_get_rate


def safe_str(value: Any) -> str:
    return "" if value is None else str(value)


def import_records_from_xlsx(
    filepath: str,
    *,
    policy: ImportPolicy,
    currency_service: Any,
    wallet_ids: set[int] | None,
    existing_initial_balance: float,
    max_import_file_size: int,
    max_import_rows: int,
    logger: Logger,
    parse_transfer_row: Callable[..., tuple[list[Record] | None, Any, int, str | None]],
    parse_import_row: Callable[..., tuple[Record | None, float | None, str | None]],
    restore_missing_transfers: Callable[[list[Record], dict[int, Any]], None],
    validate_transfer_integrity: Callable[
        [list[Record], dict[int, Any], set[int] | None], list[str]
    ],
    finalize_workbook_io: Callable[..., None],
) -> tuple[list[Record], float, ImportSummary]:
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"XLSX file not found: {filepath}")
    if os.path.getsize(filepath) > max_import_file_size:
        raise ValueError(f"XLSX file is too large: {os.path.getsize(filepath)} bytes")

    wb = load_workbook(filepath, data_only=True, read_only=True)
    try:
        if not wb.worksheets:
            return [], 0.0, (0, 0, [])

        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        first_row = next(rows_iter, None)
        if first_row is None:
            return [], 0.0, (0, 0, [])

        header_offset = 0
        header_row = first_row
        first_cell = safe_str(first_row[0]).strip()
        if is_statement_title(first_cell):
            header_row = next(rows_iter, None)
            header_offset = 1
        if header_row is None:
            return [], 0.0, (0, 0, [])

        headers = [canonical_report_header(safe_str(h)) for h in header_row]
        records: list[Record] = []
        initial_balance = to_money_float(existing_initial_balance)
        errors: list[str] = []
        skipped = 0
        imported = 0

        transfers: dict[int, Any] = {}
        next_transfer_id = 1

        is_report_xlsx = {"date", "type", "category", "amount"}.issubset(
            set(headers)
        ) and "amount_original" not in set(headers)
        get_rate = None
        if policy == ImportPolicy.CURRENT_RATE:
            get_rate = resolve_get_rate(currency_service)
        seen_rows = 0
        seen_initial_balance = False

        for idx, row in enumerate(rows_iter, start=header_offset + 2):
            seen_rows += 1
            if seen_rows > max_import_rows:
                raise ValueError(f"XLSX import exceeded row limit ({max_import_rows})")
            if not row or all(cell is None for cell in row):
                continue

            raw = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

            if is_report_xlsx:
                date_value = safe_str(raw.get("date", "")).strip()
                if is_report_total_row_label(date_value):
                    continue
                row_type_value = canonical_report_row_type(safe_str(raw.get("type", "")).strip())
                amount_value = raw.get("amount")
                if date_value == "" and row_type_value in {"initial_balance", "opening_balance"}:
                    raw["type"] = "initial_balance"
                    raw["amount_original"] = amount_value
                else:
                    raw["type"] = row_type_value or raw.get("type", "")
                    raw["amount"] = amount_value

            row_type = safe_type(safe_str(raw.get("type", "")).lower())
            if row_type == "transfer":
                row_lc = {norm_key(str(k)): str(v) if v is not None else "" for k, v in raw.items()}
                parsed_records, transfer, next_transfer_id, error = parse_transfer_row(
                    row_lc=row_lc,
                    row_label=f"row {idx}",
                    policy=policy,
                    get_rate=get_rate,
                    next_transfer_id=next_transfer_id,
                    wallet_ids=wallet_ids,
                )
                if error:
                    skipped += 1
                    errors.append(error)
                    logger.warning("XLSX import skipped %s", error)
                    continue
                if transfer is not None and parsed_records is not None:
                    if transfer.id in transfers:
                        skipped += 1
                        errors.append(f"row {idx}: duplicate transfer_id #{transfer.id}")
                        continue
                    transfers[transfer.id] = transfer
                    records.extend(parsed_records)
                    imported += 1
                continue

            record, parsed_balance, error = parse_import_row(
                raw,
                row_label=f"row {idx}",
                policy=policy,
                get_rate=get_rate,
                mandatory_only=False,
            )
            if error:
                skipped += 1
                errors.append(error)
                logger.warning("XLSX import skipped %s", error)
                continue
            if parsed_balance is not None:
                if seen_initial_balance:
                    skipped += 1
                    errors.append(f"row {idx}: duplicate initial_balance row")
                    logger.warning("XLSX import skipped row %s: duplicate initial_balance row", idx)
                    continue
                seen_initial_balance = True
                initial_balance = parsed_balance
                continue
            if record is None:
                continue
            if wallet_ids is not None and record.wallet_id not in wallet_ids:
                skipped += 1
                errors.append(f"row {idx}: wallet not found ({record.wallet_id})")
                continue
            imported += 1
            records.append(record)
            if record.transfer_id is not None:
                next_transfer_id = max(next_transfer_id, int(record.transfer_id) + 1)

        restore_missing_transfers(records, transfers)
        integrity_errors = validate_transfer_integrity(records, transfers, wallet_ids)
        if integrity_errors:
            skipped += len(integrity_errors)
            errors.extend(integrity_errors)

        logger.info(
            "XLSX import completed: imported=%s skipped=%s file=%s",
            imported,
            skipped,
            filepath,
        )
        return records, initial_balance, (imported, skipped, errors)
    finally:
        finalize_workbook_io(wb, context="records XLSX import", logger=logger)


def import_mandatory_expenses_from_xlsx(
    filepath: str,
    *,
    policy: ImportPolicy,
    currency_service: Any,
    max_import_file_size: int,
    max_import_rows: int,
    logger: Logger,
    parse_import_row: Callable[..., tuple[Record | None, float | None, str | None]],
    finalize_workbook_io: Callable[..., None],
) -> tuple[list[MandatoryExpenseRecord], ImportSummary]:
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"XLSX file not found: {filepath}")
    if os.path.getsize(filepath) > max_import_file_size:
        raise ValueError(f"XLSX file is too large: {os.path.getsize(filepath)} bytes")

    wb = load_workbook(filepath, data_only=True, read_only=True)
    try:
        if not wb.worksheets:
            return [], (0, 0, [])

        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return [], (0, 0, [])

        headers = [norm_key(safe_str(h)) for h in header_row]
        expenses: list[MandatoryExpenseRecord] = []
        errors: list[str] = []
        skipped = 0
        imported = 0

        get_rate = None
        if policy == ImportPolicy.CURRENT_RATE:
            get_rate = resolve_get_rate(currency_service)
        seen_rows = 0

        for idx, row in enumerate(rows_iter, start=2):
            seen_rows += 1
            if seen_rows > max_import_rows:
                raise ValueError(f"XLSX import exceeded row limit ({max_import_rows})")
            if not row or all(cell is None for cell in row):
                continue
            raw = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

            record, _, error = parse_import_row(
                raw,
                row_label=f"row {idx}",
                policy=policy,
                get_rate=get_rate,
                mandatory_only=True,
            )
            if error:
                skipped += 1
                errors.append(error)
                logger.warning("Mandatory XLSX import skipped %s", error)
                continue
            if isinstance(record, MandatoryExpenseRecord):
                imported += 1
                expenses.append(record)

        logger.info(
            "Mandatory XLSX import completed: imported=%s skipped=%s file=%s",
            imported,
            skipped,
            filepath,
        )
        return expenses, (imported, skipped, errors)
    finally:
        finalize_workbook_io(wb, context="mandatory XLSX import", logger=logger)
