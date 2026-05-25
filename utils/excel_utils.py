import logging

from openpyxl import Workbook

from domain.import_policy import ImportPolicy
from domain.records import MandatoryExpenseRecord, Record
from domain.reports import Report
from services.importing.parser import parse_transfer_row
from utils.csv_utils import _restore_missing_transfers, _validate_transfer_integrity
from utils.export.io import finalize_workbook_io, save_workbook_output
from utils.import_core import ImportSummary, parse_import_row
from utils.records.tabular import (
    mandatory_expense_export_rows,
    record_export_rows,
)
from utils.spreadsheets.excel_imports import (
    import_mandatory_expenses_from_xlsx as _import_mandatory_expenses_from_xlsx,
)
from utils.spreadsheets.excel_imports import import_records_from_xlsx as _import_records_from_xlsx
from utils.spreadsheets.excel_report_exports import (
    grouped_report_to_xlsx as _grouped_report_to_xlsx,
)
from utils.spreadsheets.excel_report_exports import (
    report_to_xlsx as _report_to_xlsx,
)
from utils.spreadsheets.excel_sheet_styles import (
    set_auto_width as _set_auto_width,
)
from utils.spreadsheets.excel_sheet_styles import (
    style_data_row as _style_data_row,
)
from utils.spreadsheets.excel_sheet_styles import (
    style_header_row as _style_header_row,
)

logger = logging.getLogger(__name__)
MAX_IMPORT_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
MAX_IMPORT_ROWS = 200_000

DATA_HEADERS = [
    "date",
    "type",
    "wallet_id",
    "category",
    "amount_original",
    "currency",
    "rate_at_operation",
    "amount_base",
    "description",
    "tags",
    "period",
    "transfer_id",
    "from_wallet_id",
    "to_wallet_id",
]
MANDATORY_HEADERS = [
    "type",
    "date",
    "wallet_id",
    "category",
    "amount_original",
    "currency",
    "rate_at_operation",
    "amount_base",
    "description",
    "period",
]


def _close_workbook_safely(workbook: Workbook, *, context: str) -> None:
    finalize_workbook_io(workbook, context=context, logger=logger)


def report_to_xlsx(
    report: Report,
    filepath: str,
    *,
    debts=None,
    base_currency: str = "KZT",
) -> None:
    _report_to_xlsx(report, filepath, debts=list(debts or []), base_currency=base_currency)


def grouped_report_to_xlsx(
    statement_title: str,
    grouped_rows: list[tuple[str, int, float]],
    filepath: str,
    *,
    base_currency: str = "KZT",
) -> None:
    _grouped_report_to_xlsx(
        statement_title,
        grouped_rows,
        filepath,
        base_currency=base_currency,
    )


def report_from_xlsx(filepath: str) -> Report:
    records, initial_balance, _ = import_records_from_xlsx(filepath, ImportPolicy.LEGACY)
    return Report(records, initial_balance)


def export_records_to_xlsx(
    records: list[Record],
    filepath: str,
    initial_balance: float = 0.0,
    *,
    transfers=None,
) -> None:
    del initial_balance  # legacy argument, kept for compatibility

    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Data"
        ws.append(DATA_HEADERS)
        _style_header_row(ws, 1)
        ws.freeze_panes = "A2"

    for row in record_export_rows(records, transfers=list(transfers or [])):
        if ws is not None:
            ws.append([row.get(header, "") for header in DATA_HEADERS])
            _style_data_row(ws, ws.max_row, amount_columns=(5, 7, 8))

    if ws is not None and ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:N{ws.max_row}"
        _set_auto_width(ws)

    save_workbook_output(wb, filepath, context="records XLSX export", logger=logger)


def import_records_from_xlsx(
    filepath: str,
    policy: ImportPolicy = ImportPolicy.FULL_BACKUP,
    currency_service=None,
    wallet_ids: set[int] | None = None,
    existing_initial_balance: float = 0.0,
) -> tuple[list[Record], float, ImportSummary]:
    return _import_records_from_xlsx(
        filepath,
        policy=policy,
        currency_service=currency_service,
        wallet_ids=wallet_ids,
        existing_initial_balance=existing_initial_balance,
        max_import_file_size=MAX_IMPORT_FILE_SIZE,
        max_import_rows=MAX_IMPORT_ROWS,
        logger=logger,
        parse_transfer_row=parse_transfer_row,
        parse_import_row=parse_import_row,
        restore_missing_transfers=_restore_missing_transfers,
        validate_transfer_integrity=_validate_transfer_integrity,
        finalize_workbook_io=finalize_workbook_io,
    )


def export_mandatory_expenses_to_xlsx(
    expenses: list[MandatoryExpenseRecord], filepath: str
) -> None:
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Mandatory"
        ws.append(MANDATORY_HEADERS)
        _style_header_row(ws, 1)
        ws.freeze_panes = "A2"

    for row in mandatory_expense_export_rows(expenses):
        if ws is not None:
            ws.append([row.get(header, "") for header in MANDATORY_HEADERS])
            _style_data_row(ws, ws.max_row, amount_columns=(4, 6, 7))

    if ws is not None and ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:I{ws.max_row}"
        _set_auto_width(ws)

    save_workbook_output(wb, filepath, context="mandatory XLSX export", logger=logger)


def import_mandatory_expenses_from_xlsx(
    filepath: str,
    policy: ImportPolicy = ImportPolicy.FULL_BACKUP,
    currency_service=None,
) -> tuple[list[MandatoryExpenseRecord], ImportSummary]:
    return _import_mandatory_expenses_from_xlsx(
        filepath,
        policy=policy,
        currency_service=currency_service,
        max_import_file_size=MAX_IMPORT_FILE_SIZE,
        max_import_rows=MAX_IMPORT_ROWS,
        logger=logger,
        parse_import_row=parse_import_row,
        finalize_workbook_io=finalize_workbook_io,
    )
