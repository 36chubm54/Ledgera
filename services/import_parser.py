from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from domain.import_policy import ImportPolicy
from domain.records import ExpenseRecord, IncomeRecord, Record
from domain.transfers import Transfer
from utils.backup_utils import unwrap_backup_payload
from utils.import_core import as_float, norm_key, parse_optional_strict_int
from utils.money import quantize_money, to_decimal, to_money_float, to_rate_float
from utils.tag_utils import normalize_tag_name, normalize_tag_names

MAX_IMPORT_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
MAX_IMPORT_ROWS = 200_000
MAX_CSV_FIELD_SIZE = 1_000_000


@dataclass(frozen=True)
class ParsedImportData:
    path: str
    file_type: str
    rows: list[dict[str, Any]] = field(default_factory=list)
    mandatory_rows: list[dict[str, Any]] = field(default_factory=list)
    budgets: list[dict[str, Any]] = field(default_factory=list)
    debts: list[dict[str, Any]] = field(default_factory=list)
    debt_payments: list[dict[str, Any]] = field(default_factory=list)
    assets: list[dict[str, Any]] = field(default_factory=list)
    asset_snapshots: list[dict[str, Any]] = field(default_factory=list)
    goals: list[dict[str, Any]] = field(default_factory=list)
    distribution_items: list[dict[str, Any]] = field(default_factory=list)
    distribution_subitems: list[dict[str, Any]] = field(default_factory=list)
    distribution_snapshots: list[dict[str, Any]] = field(default_factory=list)
    wallets: list[dict[str, Any]] = field(default_factory=list)
    tags: list[dict[str, Any]] = field(default_factory=list)
    record_tags: list[dict[str, Any]] = field(default_factory=list)
    initial_balance: float | None = None
    json_sections_present: frozenset[str] = field(default_factory=frozenset)


def parse_import_file(path: str, *, force: bool = False) -> ParsedImportData:
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f"Import file not found: {path}")
    if source.stat().st_size > MAX_IMPORT_FILE_SIZE:
        raise ValueError(f"Import file is too large: {source.stat().st_size} bytes")
    suffix = source.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv_rows(path)
        return ParsedImportData(path=path, file_type="csv", rows=rows)
    if suffix in {".xlsx", ".xlsm"}:
        rows = _read_xlsx_rows(path)
        return ParsedImportData(path=path, file_type="xlsx", rows=rows)
    if suffix == ".json":
        return _read_json_payload(path, force=force)
    raise ValueError(f"Unsupported import file type: {suffix}")


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    return {norm_key(str(k)): v for k, v in row.items() if k is not None}


def _validate_currency(currency: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z]{3}", currency or ""))


def parse_transfer_row(
    row_lc: dict[str, str],
    *,
    row_label: str,
    policy: ImportPolicy,
    get_rate,
    next_transfer_id: int,
    wallet_ids: set[int] | None,
) -> tuple[list[Record] | None, Transfer | None, int, str | None]:
    date_value = str(row_lc.get("date", "") or "").strip()
    if not date_value:
        return None, None, next_transfer_id, f"{row_label}: missing required field 'date'"

    from_wallet_id = parse_optional_strict_int(row_lc.get("from_wallet_id")) or 0
    to_wallet_id = parse_optional_strict_int(row_lc.get("to_wallet_id")) or 0
    if from_wallet_id <= 0 or to_wallet_id <= 0:
        return (
            None,
            None,
            next_transfer_id,
            f"{row_label}: invalid transfer wallets (from_wallet_id/to_wallet_id)",
        )
    if from_wallet_id == to_wallet_id:
        return None, None, next_transfer_id, f"{row_label}: transfer wallets must be different"

    if wallet_ids is not None:
        if from_wallet_id not in wallet_ids:
            return None, None, next_transfer_id, f"{row_label}: wallet not found ({from_wallet_id})"
        if to_wallet_id not in wallet_ids:
            return None, None, next_transfer_id, f"{row_label}: wallet not found ({to_wallet_id})"

    if policy == ImportPolicy.LEGACY:
        amount_original = as_float(row_lc.get("amount"), None)
        if amount_original is None:
            return None, None, next_transfer_id, f"{row_label}: invalid amount"
        amount_original = to_money_float(abs(to_decimal(amount_original)))
        currency = "KZT"
        rate_at_operation = 1.0
        amount_kzt = amount_original
    else:
        amount_original = as_float(row_lc.get("amount_original"), None)
        if amount_original is None:
            amount_original = as_float(row_lc.get("amount"), None)
        if amount_original is None:
            return None, None, next_transfer_id, f"{row_label}: invalid amount_original"
        amount_original = to_money_float(abs(to_decimal(amount_original)))

        currency = str(row_lc.get("currency", "KZT") or "KZT").strip().upper()
        if not _validate_currency(currency):
            return None, None, next_transfer_id, f"{row_label}: invalid currency '{currency}'"
        rate_at_operation = as_float(row_lc.get("rate_at_operation"), None)
        amount_kzt = as_float(row_lc.get("amount_kzt"), None)

    if policy == ImportPolicy.CURRENT_RATE:
        if get_rate is None:
            return (
                None,
                None,
                next_transfer_id,
                f"{row_label}: current-rate policy requires currency service",
            )
        try:
            rate_at_operation = to_rate_float(get_rate(currency))
            amount_kzt = to_money_float(
                quantize_money(amount_original) * to_decimal(rate_at_operation)
            )
        except Exception as exc:
            return (
                None,
                None,
                next_transfer_id,
                f"{row_label}: failed to get current rate for {currency} ({exc})",
            )

    if rate_at_operation is None:
        return (
            None,
            None,
            next_transfer_id,
            f"{row_label}: missing required field 'rate_at_operation'",
        )
    if amount_kzt is None:
        return None, None, next_transfer_id, f"{row_label}: missing required field 'amount_kzt'"

    transfer_id = parse_optional_strict_int(row_lc.get("transfer_id")) or 0
    if transfer_id <= 0:
        transfer_id = next_transfer_id
        next_transfer_id += 1

    description = str(row_lc.get("description", "") or "")
    category = str(row_lc.get("category", "Transfer") or "Transfer").strip() or "Transfer"

    try:
        transfer = Transfer(
            id=transfer_id,
            from_wallet_id=from_wallet_id,
            to_wallet_id=to_wallet_id,
            date=date_value,
            amount_original=amount_original,
            currency=currency,
            rate_at_operation=to_rate_float(rate_at_operation),
            amount_kzt=to_money_float(amount_kzt),
            description=description,
        )
    except Exception as exc:
        return None, None, next_transfer_id, f"{row_label}: invalid transfer ({exc})"

    expense_record = ExpenseRecord(
        date=date_value,
        wallet_id=from_wallet_id,
        transfer_id=transfer.id,
        amount_original=amount_original,
        currency=currency,
        rate_at_operation=to_rate_float(rate_at_operation),
        amount_kzt=to_money_float(amount_kzt),
        category=category,
        description=description,
    )
    income_record = IncomeRecord(
        date=date_value,
        wallet_id=to_wallet_id,
        transfer_id=transfer.id,
        amount_original=amount_original,
        currency=currency,
        rate_at_operation=to_rate_float(rate_at_operation),
        amount_kzt=to_money_float(amount_kzt),
        category=category,
        description=description,
    )
    return [expense_record, income_record], transfer, next_transfer_id, None


def _read_csv_rows(path: str) -> list[dict[str, Any]]:
    # Guardrail: prevent pathological CSV fields from allocating huge amounts of memory.
    # (This is a process-global limit in the stdlib csv module.)
    try:
        csv.field_size_limit(MAX_CSV_FIELD_SIZE)
    except (OverflowError, ValueError):
        # Extremely defensive fallback (shouldn't happen for sane MAX_CSV_FIELD_SIZE values).
        csv.field_size_limit(1_000_000)
    with open(path, newline="", encoding="utf-8") as csv_file:
        first_pos = csv_file.tell()
        first_line = ""
        while True:
            pos = csv_file.tell()
            line = csv_file.readline()
            if line == "":
                break
            if line.strip():
                first_pos = pos
                first_line = line
                break
        normalized = first_line.lstrip("\ufeff").strip()
        if not normalized.startswith("Transaction statement"):
            csv_file.seek(first_pos)
        reader = csv.DictReader(csv_file)
        rows: list[dict[str, Any]] = []
        for index, row in enumerate(reader, start=1):
            if index > MAX_IMPORT_ROWS:
                raise ValueError(f"CSV import exceeded row limit ({MAX_IMPORT_ROWS})")
            if row:
                rows.append(_normalize_row(row))
        return rows


def _read_xlsx_rows(path: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if not wb.worksheets:
            return []
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return []
        if header_row and str(header_row[0] or "").strip().startswith("Transaction statement"):
            header_row = next(rows_iter, None)
        if header_row is None:
            return []
        headers = [norm_key(str(cell or "")) for cell in header_row]
        rows: list[dict[str, Any]] = []
        for index, row in enumerate(rows_iter, start=1):
            if index > MAX_IMPORT_ROWS:
                raise ValueError(f"XLSX import exceeded row limit ({MAX_IMPORT_ROWS})")
            payload = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            if not any(str(value or "").strip() for value in payload.values()):
                continue
            rows.append(payload)
        return rows
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _read_json_payload(path: str, *, force: bool = False) -> ParsedImportData:
    with open(path, encoding="utf-8") as fp:
        payload = json.load(fp)
    payload = unwrap_backup_payload(payload, force=force)
    json_sections_present = frozenset(str(key) for key in payload.keys())

    wallets = payload.get("wallets", [])
    if not isinstance(wallets, list):
        wallets = []
    records = payload.get("records", [])
    if not isinstance(records, list):
        records = []
    mandatory = payload.get("mandatory_expenses", [])
    if not isinstance(mandatory, list):
        mandatory = []
    budgets = payload.get("budgets", [])
    if not isinstance(budgets, list):
        budgets = []
    debts = payload.get("debts", [])
    if not isinstance(debts, list):
        debts = []
    debt_payments = payload.get("debt_payments", [])
    if not isinstance(debt_payments, list):
        debt_payments = []
    assets = payload.get("assets", [])
    if not isinstance(assets, list):
        assets = []
    asset_snapshots = payload.get("asset_snapshots", [])
    if not isinstance(asset_snapshots, list):
        asset_snapshots = []
    goals = payload.get("goals", [])
    if not isinstance(goals, list):
        goals = []
    distribution_items = payload.get("distribution_items", [])
    if not isinstance(distribution_items, list):
        distribution_items = []
    distribution_subitems = payload.get("distribution_subitems", [])
    if not isinstance(distribution_subitems, list):
        distribution_subitems = []
    distribution_snapshots = payload.get("distribution_snapshots", [])
    if not isinstance(distribution_snapshots, list):
        distribution_snapshots = []
    transfers = payload.get("transfers", [])
    if not isinstance(transfers, list):
        transfers = []
    tags = payload.get("tags", [])
    if not isinstance(tags, list):
        tags = []
    record_tags = payload.get("record_tags", [])
    if not isinstance(record_tags, list):
        record_tags = []

    tags_by_id: dict[int, str] = {}
    for item in tags:
        if not isinstance(item, dict):
            continue
        tag_id = parse_optional_strict_int(item.get("id"))
        tag_name = normalize_tag_name(item.get("name"))
        if tag_id is None or tag_id <= 0 or not tag_name:
            continue
        tags_by_id[tag_id] = tag_name

    record_tag_names: dict[int, list[str]] = {}
    for item in record_tags:
        if not isinstance(item, dict):
            continue
        record_id = parse_optional_strict_int(item.get("record_id"))
        tag_id = parse_optional_strict_int(item.get("tag_id"))
        inline_name = normalize_tag_name(item.get("name"))
        if record_id is None or record_id <= 0:
            continue
        tag_name = inline_name or tags_by_id.get(int(tag_id or 0), "")
        if not tag_name:
            continue
        record_tag_names.setdefault(record_id, []).append(tag_name)

    rows = [_normalize_row(item) for item in records if isinstance(item, dict)]
    for row in rows:
        record_id = parse_optional_strict_int(row.get("id"))
        if record_id is None or record_id <= 0:
            continue
        merged = list(row.get("tags", []) or [])
        merged.extend(record_tag_names.get(record_id, []))
        row["tags"] = list(normalize_tag_names(tuple(merged)))
    existing_transfer_ids = {
        transfer_id
        for item in rows
        if (transfer_id := parse_optional_strict_int(item.get("transfer_id"))) is not None
        and transfer_id > 0
    }
    rows.extend(_transfer_rows_from_aggregates(transfers, existing_transfer_ids))
    mandatory_rows = []
    for item in mandatory:
        if not isinstance(item, dict):
            continue
        normalized = _normalize_row(item)
        if not str(normalized.get("type", "") or "").strip():
            normalized["type"] = "mandatory_expense"
        mandatory_rows.append(normalized)

    initial_balance = None
    if "initial_balance" in payload:
        initial_balance = as_float(payload.get("initial_balance"), None)
    if initial_balance is None:
        for wallet in wallets:
            if not isinstance(wallet, dict):
                continue
            wallet_id = parse_optional_strict_int(wallet.get("id")) or 0
            if wallet_id == 1 or bool(wallet.get("system", False)):
                initial_balance = float(as_float(wallet.get("initial_balance"), 0.0) or 0.0)
                break

    return ParsedImportData(
        path=path,
        file_type="json",
        rows=rows,
        mandatory_rows=mandatory_rows,
        budgets=[item for item in budgets if isinstance(item, dict)],
        debts=[item for item in debts if isinstance(item, dict)],
        debt_payments=[item for item in debt_payments if isinstance(item, dict)],
        assets=[item for item in assets if isinstance(item, dict)],
        asset_snapshots=[item for item in asset_snapshots if isinstance(item, dict)],
        goals=[item for item in goals if isinstance(item, dict)],
        distribution_items=[item for item in distribution_items if isinstance(item, dict)],
        distribution_subitems=[item for item in distribution_subitems if isinstance(item, dict)],
        distribution_snapshots=[item for item in distribution_snapshots if isinstance(item, dict)],
        wallets=[wallet for wallet in wallets if isinstance(wallet, dict)],
        tags=[item for item in tags if isinstance(item, dict)],
        record_tags=[item for item in record_tags if isinstance(item, dict)],
        initial_balance=initial_balance,
        json_sections_present=json_sections_present,
    )


def _transfer_rows_from_aggregates(
    items: list[Any], existing_transfer_ids: set[int]
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        transfer_id = parse_optional_strict_int(item.get("id")) or 0
        if transfer_id > 0 and transfer_id in existing_transfer_ids:
            continue
        rows.append(
            {
                "type": "transfer",
                "date": item.get("date", ""),
                "description": item.get("description", ""),
                "transfer_id": transfer_id if transfer_id > 0 else item.get("id"),
                "amount_original": item.get("amount_original"),
                "currency": item.get("currency", "KZT"),
                "rate_at_operation": item.get("rate_at_operation"),
                "amount_kzt": item.get("amount_kzt"),
                "from_wallet_id": item.get("from_wallet_id"),
                "to_wallet_id": item.get("to_wallet_id"),
            }
        )
    return rows
