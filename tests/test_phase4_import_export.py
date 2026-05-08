import csv
import json
import os
import tempfile
import time

from openpyxl import load_workbook

from domain.import_policy import ImportPolicy
from domain.records import ExpenseRecord, IncomeRecord, MandatoryExpenseRecord
from domain.transfers import Transfer
from infrastructure.repositories import JsonFileRecordRepository
from utils.backup_utils import import_full_backup_from_json
from utils.csv_utils import DATA_HEADERS, export_records_to_csv, import_records_from_csv
from utils.excel_utils import export_records_to_xlsx, import_records_from_xlsx


def test_json_import_new_structure():
    payload = {
        "wallets": [
            {
                "id": 1,
                "name": "Main wallet",
                "currency": "KZT",
                "initial_balance": 100.0,
                "system": True,
                "allow_negative": False,
                "is_active": True,
            },
            {
                "id": 2,
                "name": "Cash",
                "currency": "KZT",
                "initial_balance": 0.0,
                "system": False,
                "allow_negative": False,
                "is_active": True,
            },
        ],
        "records": [
            {
                "date": "2026-01-01",
                "type": "expense",
                "wallet_id": 1,
                "transfer_id": 10,
                "category": "Transfer",
                "amount_original": 50,
                "currency": "KZT",
                "rate_at_operation": 1,
                "amount_kzt": 50,
                "description": "move",
            },
            {
                "date": "2026-01-01",
                "type": "income",
                "wallet_id": 2,
                "transfer_id": 10,
                "category": "Transfer",
                "amount_original": 50,
                "currency": "KZT",
                "rate_at_operation": 1,
                "amount_kzt": 50,
                "description": "move",
            },
        ],
        "mandatory_expenses": [],
        "transfers": [
            {
                "id": 10,
                "from_wallet_id": 1,
                "to_wallet_id": 2,
                "date": "2026-01-01",
                "amount_original": 50,
                "currency": "KZT",
                "rate_at_operation": 1,
                "amount_kzt": 50,
                "description": "move",
            }
        ],
    }

    with tempfile.NamedTemporaryFile(
        mode="w", delete=False, suffix=".json", encoding="utf-8"
    ) as fp:
        json.dump(payload, fp)
        path = fp.name
    try:
        result = import_full_backup_from_json(path)
        wallets = result.wallets
        records = result.records
        mandatory = result.mandatory_expenses
        transfers = result.transfers
        summary = result.summary
        assert len(wallets) == 2
        assert len(records) == 2
        assert mandatory == []
        assert len(transfers) == 1
        assert summary[1] == 0
    finally:
        os.unlink(path)


def test_csv_import_transfer_one_row_restores_two_records():
    with tempfile.NamedTemporaryFile(
        mode="w", delete=False, suffix=".csv", newline="", encoding="utf-8"
    ) as fp:
        writer = csv.DictWriter(fp, fieldnames=DATA_HEADERS)
        writer.writeheader()
        writer.writerow(
            {
                "date": "2026-01-01",
                "type": "transfer",
                "wallet_id": "",
                "category": "Transfer",
                "amount_original": 25,
                "currency": "KZT",
                "rate_at_operation": 1,
                "amount_kzt": 25,
                "description": "move",
                "period": "",
                "transfer_id": 77,
                "from_wallet_id": 1,
                "to_wallet_id": 2,
            }
        )
        path = fp.name
    try:
        records, _, summary = import_records_from_csv(
            path,
            policy=ImportPolicy.FULL_BACKUP,
            wallet_ids={1, 2},
        )
        transfer_records = [r for r in records if r.transfer_id == 77]
        assert summary[1] == 0
        assert len(transfer_records) == 2
        assert {r.type for r in transfer_records} == {"expense", "income"}
    finally:
        os.unlink(path)


def test_xlsx_import_transfer_one_row_restores_two_records():
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as fp:
        path = fp.name
    try:
        wb_rows = [
            DATA_HEADERS,
            [
                "2026-01-01",
                "transfer",
                "",
                "Transfer",
                25,
                "KZT",
                1,
                25,
                "move",
                "",
                "",
                78,
                1,
                2,
            ],
        ]
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Data"
            for row in wb_rows:
                ws.append(row)
        wb.save(path)
        wb.close()

        records, _, summary = import_records_from_xlsx(
            path,
            policy=ImportPolicy.FULL_BACKUP,
            wallet_ids={1, 2},
        )
        transfer_records = [r for r in records if r.transfer_id == 78]
        assert summary[1] == 0
        assert len(transfer_records) == 2
        assert {r.type for r in transfer_records} == {"expense", "income"}
    finally:
        os.unlink(path)


def test_export_csv_xlsx_transfer_one_row():
    records = [
        IncomeRecord(
            date="2026-01-02",
            wallet_id=1,
            amount_original=10,
            currency="KZT",
            rate_at_operation=1,
            amount_kzt=10,
            category="Salary",
        ),
        ExpenseRecord(
            date="2026-01-01",
            wallet_id=1,
            transfer_id=9,
            amount_original=30,
            currency="KZT",
            rate_at_operation=1,
            amount_kzt=30,
            category="Transfer",
        ),
        IncomeRecord(
            date="2026-01-01",
            wallet_id=2,
            transfer_id=9,
            amount_original=30,
            currency="KZT",
            rate_at_operation=1,
            amount_kzt=30,
            category="Transfer",
        ),
        MandatoryExpenseRecord(
            date="2026-01-03",
            wallet_id=1,
            amount_original=5,
            currency="KZT",
            rate_at_operation=1,
            amount_kzt=5,
            category="Rent",
            description="monthly",
            period="monthly",
        ),
    ]
    transfers = [
        Transfer(
            id=9,
            from_wallet_id=1,
            to_wallet_id=2,
            date="2026-01-01",
            amount_original=30,
            currency="KZT",
            rate_at_operation=1,
            amount_kzt=30,
            description="move",
        )
    ]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as csv_fp:
        csv_path = csv_fp.name
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as xlsx_fp:
        xlsx_path = xlsx_fp.name
    try:
        export_records_to_csv(records, csv_path, transfers=transfers)
        with open(csv_path, encoding="utf-8") as fp:
            text = fp.read()
        assert "initial_balance" not in text
        assert ",transfer," in text

        export_records_to_xlsx(records, xlsx_path, transfers=transfers)
        wb = load_workbook(xlsx_path, data_only=True)
        try:
            ws = wb.worksheets[0]
            rows = list(ws.iter_rows(values_only=True))
            types = [str(row[1]) for row in rows[1:] if row and row[1] is not None]
            assert "transfer" in types
            assert "initial_balance" not in types
        finally:
            wb.close()
    finally:
        os.unlink(csv_path)
        for _ in range(5):
            try:
                os.unlink(xlsx_path)
                break
            except PermissionError:
                time.sleep(0.1)


def test_repository_does_not_store_global_initial_balance_key():
    with tempfile.NamedTemporaryFile(delete=False, suffix=".json") as fp:
        path = fp.name
    try:
        os.unlink(path)
        repo = JsonFileRecordRepository(path)
        repo.save_initial_balance(55.0)
        with open(path, encoding="utf-8") as raw:
            payload = json.load(raw)
        assert "initial_balance" not in payload
        assert payload["wallets"][0]["initial_balance"] == 55.0
    finally:
        os.unlink(path)


def test_json_import_old_format_migrates_global_balance_to_main_wallet():
    legacy_payload = {
        "initial_balance": 200.0,
        "records": [
            {
                "date": "2026-01-01",
                "type": "income",
                "category": "Salary",
                "amount_original": 10,
                "currency": "KZT",
                "rate_at_operation": 1,
                "amount_kzt": 10,
            }
        ],
        "mandatory_expenses": [],
    }

    with tempfile.NamedTemporaryFile(
        mode="w", delete=False, suffix=".json", encoding="utf-8"
    ) as fp:
        json.dump(legacy_payload, fp)
        path = fp.name
    try:
        result = import_full_backup_from_json(path)
        wallets = result.wallets
        records = result.records
        transfers = result.transfers
        summary = result.summary
        assert summary[1] == 0
        assert wallets[0].id == 1
        assert wallets[0].initial_balance == 200.0
        assert len(records) == 1
        assert transfers == []
    finally:
        os.unlink(path)
