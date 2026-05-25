import logging
import os
import tempfile
import time
from datetime import date as dt_date

import pytest
from openpyxl import load_workbook

import domain.reports as reports_module
from domain.debt import Debt, DebtKind, DebtStatus
from domain.import_policy import ImportPolicy
from domain.records import (
    ExpenseRecord,
    IncomeRecord,
    MandatoryExpenseRecord,
)
from domain.reports import Report
from domain.transfers import Transfer
from gui.i18n import get_language, set_language
from utils.csv_utils import (
    export_mandatory_expenses_to_csv,
    import_mandatory_expenses_from_csv,
)
from utils.excel_utils import (
    _close_workbook_safely,
    export_mandatory_expenses_to_xlsx,
    export_records_to_xlsx,
    import_mandatory_expenses_from_xlsx,
    import_records_from_xlsx,
    report_from_xlsx,
    report_to_xlsx,
)


@pytest.fixture(autouse=True)
def _english_report_exports():
    previous = get_language()
    set_language("en")
    try:
        yield
    finally:
        set_language(previous)


def test_report_xlsx_roundtrip():
    records = [
        IncomeRecord(date="2025-01-01", _amount_init=100.0, category="Salary", tags=("work",)),
        ExpenseRecord(date="2025-01-02", _amount_init=30.0, category="Food", tags=("food", "home")),
    ]
    report = Report(records, initial_balance=50.0)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        report_to_xlsx(report, tmp_path)
        wb = load_workbook(tmp_path, data_only=True)
        try:
            report_ws = wb["Report"]
            assert report_ws.cell(1, 1).value == "Transaction statement"
            assert report_ws.cell(4, 2).value == "Initial balance"
            assert report_ws.cell(4, 4).value == 50.0
            assert report_ws.cell(5, 1).value == "2025-01-02"
            assert report_ws.cell(5, 2).value == "Expense"
            assert report_ws.cell(6, 1).value == "2025-01-01"
            assert report_ws.cell(6, 2).value == "Income"
            assert report_ws.freeze_panes == "A3"
            assert report_ws.auto_filter.ref == "A2:E8"
            assert "Yearly Report" in wb.sheetnames
            assert "By Category" in wb.sheetnames
            assert "By Tag" in wb.sheetnames
            assert wb.sheetnames[:4] == ["Report", "By Category", "By Tag", "Yearly Report"]
            summary_ws = wb["Yearly Report"]
            assert summary_ws.cell(1, 1).value == "Month (2025)"
            assert summary_ws.freeze_panes == "A2"
            assert summary_ws.cell(2, 2).value == 100.0
            bycat_ws = wb["By Category"]
            labels = [row[0] for row in bycat_ws.iter_rows(max_col=1, values_only=True) if row[0]]
            assert "Category: Food" in labels
            assert "Category: Salary" in labels
            bytag_ws = wb["By Tag"]
            tag_labels = [
                row[0] for row in bytag_ws.iter_rows(max_col=1, values_only=True) if row[0]
            ]
            assert "Tag: #food" in tag_labels
            assert "Tag: #home" in tag_labels
            assert "Tag: #work" in tag_labels
        finally:
            wb.close()
        imported = report_from_xlsx(tmp_path)
        assert len(imported.records()) == 2
        assert abs(imported._initial_balance - 50.0) < 1e-6
        assert abs(imported.total() - report.total()) < 1e-6
    finally:
        os.unlink(tmp_path)


def test_close_workbook_safely_logs_expected_close_failures(caplog) -> None:
    class _BrokenWorkbook:
        def close(self) -> None:
            raise RuntimeError("close failed")

    caplog.set_level(logging.DEBUG)

    _close_workbook_safely(_BrokenWorkbook(), context="test export")  # type: ignore[arg-type]

    assert "Workbook close degraded after test export" in caplog.text
    assert "close failed" in caplog.text


def test_report_xlsx_by_tag_sheet_repeats_multi_tag_record_under_each_single_tag():
    report = Report(
        [
            ExpenseRecord(
                date="2025-01-02",
                _amount_init=30.0,
                category="Food",
                tags=("food", "home"),
            ),
            ExpenseRecord(
                date="2025-01-03",
                _amount_init=15.0,
                category="Home",
                tags=("home",),
            ),
        ],
        initial_balance=0.0,
    )

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        report_to_xlsx(report, tmp_path)
        wb = load_workbook(tmp_path, data_only=True)
        try:
            ws = wb["By Tag"]
            rows = [row for row in ws.iter_rows(values_only=True)]
            assert ("Tag: #food", None, None, None, None) in rows
            assert ("Tag: #home", None, None, None, None) in rows

            food_idx = rows.index(("Tag: #food", None, None, None, None))
            home_idx = rows.index(("Tag: #home", None, None, None, None))
            food_block = rows[food_idx:home_idx]
            home_block = rows[home_idx:]

            assert food_block[2] == ("2025-01-02", "Expense", "Food", 30.0, "#food #home")
            assert home_block[2] == ("2025-01-03", "Expense", "Home", 15.0, "#home")
            assert home_block[3] == ("2025-01-02", "Expense", "Food", 30.0, "#food #home")
        finally:
            wb.close()
    finally:
        os.unlink(tmp_path)


def test_report_xlsx_includes_debts_sheet_when_provided():
    records = [
        IncomeRecord(date="2025-01-01", _amount_init=100.0, category="Salary"),
        ExpenseRecord(date="2025-01-02", _amount_init=30.0, category="Food"),
    ]
    report = Report(records, initial_balance=50.0)
    debts = [
        Debt(
            id=1,
            contact_name="Alex",
            kind=DebtKind.DEBT,
            total_amount_minor=10000,
            remaining_amount_minor=7500,
            currency="KZT",
            interest_rate=0.0,
            status=DebtStatus.OPEN,
            created_at="2026-03-01",
        )
    ]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        report_to_xlsx(report, tmp_path, debts=debts)
        wb = load_workbook(tmp_path, data_only=True)
        try:
            assert "Debts" in wb.sheetnames
            ws = wb["Debts"]
            assert ws.cell(1, 1).value == "Contact"
            assert ws.cell(2, 1).value == "Alex"
            assert ws.cell(1, 4).value == "Opened"
            assert ws.cell(1, 9).value == "Settled"
            assert ws.cell(1, 10).value == "Progress %"
            assert ws.cell(2, 4).value == "2026-03-01"
            assert ws.cell(2, 7).value == 100.0
            assert ws.cell(2, 8).value == 75.0
            assert ws.cell(2, 9).value == 25.0
            assert ws.cell(2, 10).value == 25.0
        finally:
            wb.close()
    finally:
        os.unlink(tmp_path)


def test_report_xlsx_skips_debts_sheet_when_no_debt_overlaps_report_period(monkeypatch):
    class FakeDate(dt_date):
        @classmethod
        def today(cls):
            return cls(2026, 4, 15)

    monkeypatch.setattr(reports_module, "dt_date", FakeDate)

    report = Report(
        [IncomeRecord(date="2026-01-01", _amount_init=100.0, category="Salary")],
        initial_balance=50.0,
    ).filter_by_period_range("2026-04")
    debts = [
        Debt(
            id=1,
            contact_name="Alex",
            kind=DebtKind.DEBT,
            total_amount_minor=10000,
            remaining_amount_minor=7500,
            currency="KZT",
            interest_rate=0.0,
            status=DebtStatus.OPEN,
            created_at="2026-01-10",
            closed_at="2026-02-05",
        )
    ]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        report_to_xlsx(report, tmp_path, debts=debts)
        wb = load_workbook(tmp_path, data_only=True)
        try:
            assert "Debts" not in wb.sheetnames
        finally:
            wb.close()
    finally:
        os.unlink(tmp_path)


def test_report_xlsx_uses_opening_balance_label_for_filtered_report():
    records = [
        IncomeRecord(date="2024-12-31", _amount_init=20.0, category="Old"),
        IncomeRecord(date="2025-01-01", _amount_init=100.0, category="Salary"),
        ExpenseRecord(date="2025-01-02", _amount_init=30.0, category="Food"),
    ]
    report = Report(records, initial_balance=50.0).filter_by_period("2025")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        report_to_xlsx(report, tmp_path)
        wb = load_workbook(tmp_path, data_only=True)
        try:
            ws = wb["Report"]
            assert ws.cell(1, 1).value == "Transaction statement (2025-01-01 - 2025-12-31)"
            assert ws.cell(4, 2).value == "Opening balance"
            assert ws.cell(4, 4).value == 70.0
        finally:
            wb.close()
    finally:
        for _ in range(5):
            try:
                os.unlink(tmp_path)
                break
            except PermissionError:
                time.sleep(0.1)


def test_report_xlsx_localizes_visible_sheets_and_keeps_roundtrip_in_ru():
    records = [
        IncomeRecord(date="2025-01-01", _amount_init=100.0, category="Зарплата", tags=("работа",)),
        ExpenseRecord(date="2025-01-02", _amount_init=30.0, category="Еда", tags=("дом",)),
    ]
    report = Report(records, initial_balance=50.0)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    previous = get_language()
    try:
        set_language("ru")
        report_to_xlsx(report, tmp_path)
        wb = load_workbook(tmp_path, data_only=True)
        try:
            assert "Отчет" in wb.sheetnames
            ws = wb["Отчет"]
            assert ws.cell(1, 1).value == "Отчет по операциям"
            assert ws.cell(2, 1).value == "Дата"
            assert ws.cell(2, 4).value == "Сумма (KZT)"
            assert ws.cell(4, 2).value == "Начальный баланс"
            assert ws.cell(5, 1).value == "2025-01-02"
            assert ws.cell(6, 1).value == "2025-01-01"
            assert "Годовой отчет" in wb.sheetnames
            assert "По категориям" in wb.sheetnames
            assert "По тегам" in wb.sheetnames
        finally:
            wb.close()
        imported = report_from_xlsx(tmp_path)
        assert len(imported.records()) == 2
        assert abs(imported.initial_balance - 50.0) < 1e-6
        assert abs(imported.total() - report.total()) < 1e-6
    finally:
        set_language(previous)
        for _ in range(5):
            try:
                os.unlink(tmp_path)
                break
            except PermissionError:
                time.sleep(0.1)


def test_report_xlsx_localized_roundtrip_after_language_switch():
    records = [
        IncomeRecord(date="2025-01-01", _amount_init=100.0, category="Зарплата", tags=("работа",)),
        ExpenseRecord(date="2025-01-02", _amount_init=30.0, category="Еда", tags=("дом",)),
    ]
    report = Report(records, initial_balance=50.0)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    previous = get_language()
    try:
        set_language("ru")
        report_to_xlsx(report, tmp_path)
        set_language("en")
        imported = report_from_xlsx(tmp_path)
        assert len(imported.records()) == 2
        assert abs(imported.initial_balance - 50.0) < 1e-6
        assert abs(imported.total() - report.total()) < 1e-6
    finally:
        set_language(previous)
        for _ in range(5):
            try:
                os.unlink(tmp_path)
                break
            except PermissionError:
                time.sleep(0.1)


def test_mandatory_xlsx_roundtrip():
    expenses = [
        MandatoryExpenseRecord(
            date="2026-03-09",
            wallet_id=7,
            _amount_init=10.0,
            category="Sub",
            description="d1",
            period="monthly",
            auto_pay=True,
        ),
        MandatoryExpenseRecord(
            date="",
            wallet_id=9,
            _amount_init=20.5,
            category="Svc",
            description="d2",
            period="yearly",
        ),
    ]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        export_mandatory_expenses_to_xlsx(expenses, tmp_path)
        imported, _ = import_mandatory_expenses_from_xlsx(tmp_path)
        assert len(imported) == 2
        assert imported[0].amount == 10.0
        assert str(imported[0].date) == "2026-03-09"
        assert imported[0].wallet_id == 7
        assert imported[0].auto_pay is True
        assert imported[1].wallet_id == 9
        assert imported[1].period == "yearly"
    finally:
        os.unlink(tmp_path)


def test_mandatory_csv_roundtrip():
    expenses = [
        MandatoryExpenseRecord(
            date="2026-03-21",
            wallet_id=7,
            _amount_init=5.0,
            category="A",
            description="x",
            period="daily",
            auto_pay=True,
        ),
    ]
    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv", mode="w", newline="") as tmp:
        tmp_path = tmp.name
    try:
        export_mandatory_expenses_to_csv(expenses, tmp_path)
        imported, _ = import_mandatory_expenses_from_csv(tmp_path)
        assert len(imported) == 1
        assert imported[0].amount == 5.0
        assert str(imported[0].date) == "2026-03-21"
        assert imported[0].wallet_id == 7
        assert imported[0].auto_pay is True
        assert imported[0].period == "daily"
    finally:
        os.unlink(tmp_path)


def test_xlsx_export_grouped_drill_down():
    """Export a category‑filtered report to XLSX (simulating grouped drill‑down)."""
    import os
    import tempfile

    from openpyxl import load_workbook

    from domain.records import ExpenseRecord, IncomeRecord
    from domain.reports import Report
    from utils.excel_utils import report_to_xlsx

    records = [
        IncomeRecord(date="2025-01-01", _amount_init=100.0, category="Salary"),
        ExpenseRecord(date="2025-01-02", _amount_init=30.0, category="Food"),
        IncomeRecord(date="2025-01-03", _amount_init=50.0, category="Salary"),
        ExpenseRecord(date="2025-01-04", _amount_init=20.0, category="Food"),
    ]
    report = Report(records, initial_balance=200.0)
    filtered = report.filter_by_category("Salary")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        report_to_xlsx(filtered, tmp_path)
        wb = load_workbook(tmp_path, data_only=True)
        try:
            ws = wb["Report"]
            # Check title
            assert ws.cell(1, 1).value == "Transaction statement"
            # Check that there is no initial balance row (because filter_by_category resets it)
            # The first data row should be the first Salary record
            data_rows = []
            for row in ws.iter_rows(min_row=4, max_col=4, values_only=True):
                if row[0] and isinstance(row[0], str) and row[0].startswith("2025"):
                    data_rows.append(row)
            assert len(data_rows) == 2
            assert data_rows[0][2] == "Salary"
            assert data_rows[0][3] == 50.0
            assert data_rows[1][2] == "Salary"
            assert data_rows[1][3] == 100.0
            # Find SUBTOTAL row
            subtotal_found = False
            for row in ws.iter_rows(min_row=1, max_col=4, values_only=True):
                if row[0] == "SUBTOTAL":
                    assert row[3] == 150.0
                    subtotal_found = True
                    break
            assert subtotal_found
            # Find FINAL BALANCE row
            final_found = False
            for row in ws.iter_rows(min_row=1, max_col=4, values_only=True):
                if row[0] == "FINAL BALANCE":
                    assert row[3] == 150.0
                    final_found = True
                    break
            assert final_found
            assert "By Category" not in wb.sheetnames
        finally:
            wb.close()
    finally:
        for _ in range(5):
            try:
                os.unlink(tmp_path)
                break
            except PermissionError:
                time.sleep(0.1)


def test_records_xlsx_export_applies_readability_styles():
    records = [
        IncomeRecord(date="2025-01-01", _amount_init=100.0, category="Salary"),
        ExpenseRecord(date="2025-01-02", _amount_init=30.0, category="Food"),
    ]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        export_records_to_xlsx(records, tmp_path)
        wb = load_workbook(tmp_path, data_only=True)
        try:
            ws = wb["Data"]
            assert ws.freeze_panes == "A2"
            assert ws.auto_filter.ref == "A1:N3"
            assert ws.cell(1, 1).value == "date"
            assert ws.cell(2, 5).value == 100.0
            assert ws.cell(2, 8).value == 100.0
            assert ws.column_dimensions["A"].width >= 12
        finally:
            wb.close()
    finally:
        for _ in range(5):
            try:
                os.unlink(tmp_path)
                break
            except PermissionError:
                time.sleep(0.1)


def test_report_xlsx_surfaces_grouped_section_warning(monkeypatch):
    report = Report(
        [IncomeRecord(date="2025-01-01", _amount_init=100.0, category="Salary")],
        initial_balance=25.0,
    )

    def _boom(_self):
        raise RuntimeError("grouping unavailable")

    monkeypatch.setattr(Report, "grouped_by_category", _boom)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        report_to_xlsx(report, tmp_path)
        wb = load_workbook(tmp_path, data_only=True)
        try:
            assert "Warnings" in wb.sheetnames
            ws = wb["Warnings"]
            assert ws.cell(1, 1).value == "Warning"
            assert "grouping unavailable" in str(ws.cell(2, 1).value)
        finally:
            wb.close()
    finally:
        for _ in range(5):
            try:
                os.unlink(tmp_path)
                break
            except PermissionError:
                time.sleep(0.1)


def test_import_records_from_xlsx_quantizes_existing_initial_balance():
    from openpyxl import Workbook

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Data"
            ws.append(["date", "type", "amount"])
        wb.save(tmp_path)
        wb.close()

        records, initial_balance, summary = import_records_from_xlsx(
            tmp_path,
            ImportPolicy.LEGACY,
            existing_initial_balance=1.005,
        )

        assert records == []
        assert initial_balance == 1.01
        assert summary == (0, 0, [])
    finally:
        os.unlink(tmp_path)


def test_export_records_to_xlsx_preserves_record_positions_with_transfers() -> None:
    records = [
        ExpenseRecord(
            id=1,
            date="2026-03-05",
            wallet_id=1,
            amount_original=50.0,
            currency="KZT",
            rate_at_operation=1.0,
            amount_base=50.0,
            category="Food",
        ),
        ExpenseRecord(
            id=2,
            date="2026-03-01",
            wallet_id=1,
            transfer_id=7,
            amount_original=25.0,
            currency="KZT",
            rate_at_operation=1.0,
            amount_base=25.0,
            category="Transfer",
            description="Move",
        ),
        IncomeRecord(
            id=3,
            date="2026-03-01",
            wallet_id=2,
            transfer_id=7,
            amount_original=25.0,
            currency="KZT",
            rate_at_operation=1.0,
            amount_base=25.0,
            category="Transfer",
            description="Move",
        ),
        IncomeRecord(
            id=4,
            date="2026-03-10",
            wallet_id=2,
            amount_original=100.0,
            currency="KZT",
            rate_at_operation=1.0,
            amount_base=100.0,
            category="Salary",
        ),
    ]
    transfers = [
        Transfer(
            id=7,
            from_wallet_id=1,
            to_wallet_id=2,
            date="2026-03-01",
            amount_original=25.0,
            currency="KZT",
            rate_at_operation=1.0,
            amount_base=25.0,
            description="Move",
        )
    ]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    try:
        export_records_to_xlsx(records, tmp_path, transfers=transfers)
        wb = load_workbook(tmp_path, data_only=True)
        try:
            ws = wb["Data"]
            exported_rows = [
                [cell for cell in row[:4]]
                for row in ws.iter_rows(min_row=2, max_col=4, values_only=True)
                if any(cell is not None for cell in row)
            ]
            assert exported_rows == [
                ["2026-03-05", "expense", 1, "Food"],
                ["2026-03-01", "transfer", None, "Transfer"],
                ["2026-03-10", "income", 2, "Salary"],
            ]
        finally:
            wb.close()
    finally:
        for _ in range(5):
            try:
                os.unlink(tmp_path)
                break
            except PermissionError:
                time.sleep(0.1)
