import json
import os
import tempfile
import time
from builtins import __import__ as builtin_import
from pathlib import Path

import pytest
from openpyxl import load_workbook

from domain.debt import Debt, DebtKind, DebtOperationType, DebtPayment, DebtStatus
from domain.records import ExpenseRecord, IncomeRecord, MandatoryExpenseRecord
from domain.reports import Report
from domain.transfers import Transfer
from domain.wallets import Wallet
from gui import exporters, importers


def make_sample_report():
    records = [
        IncomeRecord(date="2025-01-01", _amount_init=100.0, category="Salary"),
        ExpenseRecord(date="2025-01-02", _amount_init=30.0, category="Food"),
        MandatoryExpenseRecord(
            date="2025-01-03",
            _amount_init=20.0,
            category="Rent",
            description="Monthly rent",
            period="monthly",
        ),
    ]
    return Report(records, initial_balance=10.0)


def test_export_report_csv_xlsx_pdf():
    report = make_sample_report()

    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as csv_tmp:
        csv_path = Path(csv_tmp.name)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as xlsx_tmp:
        xlsx_path = Path(xlsx_tmp.name)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as pdf_tmp:
        pdf_path = Path(pdf_tmp.name)

    try:
        exporters.export_report(report, str(csv_path), "csv")
        assert csv_path.exists()
        text = csv_path.read_text(encoding="utf-8")
        assert "Date" in text
        assert "FINAL BALANCE" in text

        exporters.export_report(report, str(xlsx_path), "xlsx")
        assert xlsx_path.exists()
        assert xlsx_path.stat().st_size > 0

        exporters.export_report(report, str(pdf_path), "pdf")
        assert pdf_path.exists()
        assert pdf_path.stat().st_size > 0
    finally:
        for p in (csv_path, xlsx_path, pdf_path):
            if p.exists():
                os.unlink(p)


def test_export_report_pdf_requires_optional_reportlab_dependency(monkeypatch):
    report = make_sample_report()

    def fake_import(name, globals=None, locals=None, fromlist=(), level=0):
        if name.startswith("reportlab"):
            raise ModuleNotFoundError("No module named 'reportlab'", name="reportlab")
        return builtin_import(name, globals, locals, fromlist, level)

    monkeypatch.delitem(__import__("sys").modules, "utils.pdf_utils", raising=False)
    monkeypatch.setattr("builtins.__import__", fake_import)

    with pytest.raises(RuntimeError, match="optional 'pdf' dependency"):
        exporters.export_report(report, "ignored.pdf", "pdf")


def test_export_and_import_mandatory_expenses_csv_xlsx():
    expenses = [
        MandatoryExpenseRecord(
            date="2026-03-15",
            wallet_id=7,
            _amount_init=50.0,
            category="Utilities",
            description="Water",
            period="monthly",
            auto_pay=True,
        ),
        MandatoryExpenseRecord(
            date="",
            wallet_id=9,
            _amount_init=120.0,
            category="Internet",
            description="ISP",
            period="monthly",
        ),
    ]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as csv_tmp:
        csv_path = Path(csv_tmp.name)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as xlsx_tmp:
        xlsx_path = Path(xlsx_tmp.name)

    try:
        exporters.export_mandatory_expenses(expenses, str(csv_path), "csv")
        assert csv_path.exists()
        data, _ = importers.import_mandatory_expenses_from_csv(str(csv_path))
        assert len(data) == len(expenses)
        assert str(data[0].date) == "2026-03-15"
        assert data[0].wallet_id == 7
        assert data[1].wallet_id == 9
        assert data[0].auto_pay is True

        exporters.export_mandatory_expenses(expenses, str(xlsx_path), "xlsx")
        assert xlsx_path.exists()
        data2, _ = importers.import_mandatory_expenses_from_xlsx(str(xlsx_path))
        assert len(data2) == len(expenses)
        assert str(data2[0].date) == "2026-03-15"
        assert data2[0].wallet_id == 7
        assert data2[1].wallet_id == 9
        assert data2[0].auto_pay is True
    finally:
        for p in (csv_path, xlsx_path):
            if p.exists():
                os.unlink(p)


def test_import_records_from_csv_xlsx_roundtrip():
    report = make_sample_report()
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as xlsx_tmp:
        xlsx_path = Path(xlsx_tmp.name)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as csv_tmp:
        csv_path = Path(csv_tmp.name)

    try:
        from utils.excel_utils import export_records_to_xlsx

        export_records_to_xlsx(report.records(), str(xlsx_path), report.initial_balance)
        records_xlsx, initial_balance_xlsx, _ = importers.import_records_from_xlsx(str(xlsx_path))
        assert len(records_xlsx) == len(report.records())
        assert initial_balance_xlsx == 0.0

        from utils.csv_utils import export_records_to_csv

        export_records_to_csv(report.records(), str(csv_path), report.initial_balance)
        records_csv, initial_balance_csv, _ = importers.import_records_from_csv(str(csv_path))
        assert len(records_csv) == len(report.records())
        assert initial_balance_csv == 0.0
    finally:
        for p in (csv_path, xlsx_path):
            if p.exists():
                for _ in range(5):
                    try:
                        os.unlink(p)
                        break
                    except PermissionError:
                        time.sleep(0.1)


def test_export_records_pipeline_preserves_order_for_csv_and_xlsx():
    records = [
        ExpenseRecord(
            id=1,
            date="2026-03-05",
            wallet_id=1,
            amount_original=50.0,
            currency="KZT",
            rate_at_operation=1.0,
            amount_base=50.0,
            category="Food",
        ),
        ExpenseRecord(
            id=2,
            date="2026-03-01",
            wallet_id=1,
            transfer_id=7,
            amount_original=25.0,
            currency="KZT",
            rate_at_operation=1.0,
            amount_base=25.0,
            category="Transfer",
            description="Move",
        ),
        IncomeRecord(
            id=3,
            date="2026-03-01",
            wallet_id=2,
            transfer_id=7,
            amount_original=25.0,
            currency="KZT",
            rate_at_operation=1.0,
            amount_base=25.0,
            category="Transfer",
            description="Move",
        ),
        IncomeRecord(
            id=4,
            date="2026-03-10",
            wallet_id=2,
            amount_original=100.0,
            currency="KZT",
            rate_at_operation=1.0,
            amount_base=100.0,
            category="Salary",
        ),
    ]
    transfers = [
        Transfer(
            id=7,
            from_wallet_id=1,
            to_wallet_id=2,
            date="2026-03-01",
            amount_original=25.0,
            currency="KZT",
            rate_at_operation=1.0,
            amount_base=25.0,
            description="Move",
        )
    ]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as csv_tmp:
        csv_path = Path(csv_tmp.name)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as xlsx_tmp:
        xlsx_path = Path(xlsx_tmp.name)

    try:
        exporters.export_records(records, str(csv_path), "csv", transfers=transfers)
        import csv

        with csv_path.open(encoding="utf-8", newline="") as handle:
            csv_rows = list(csv.DictReader(handle))
        assert [row["type"] for row in csv_rows] == ["expense", "transfer", "income"]
        assert [row["date"] for row in csv_rows] == ["2026-03-05", "2026-03-01", "2026-03-10"]

        exporters.export_records(records, str(xlsx_path), "xlsx", transfers=transfers)
        wb = load_workbook(xlsx_path, data_only=True)
        try:
            ws = wb["Data"]
            xlsx_rows = [
                list(row[:4])
                for row in ws.iter_rows(min_row=2, max_col=4, values_only=True)
                if any(cell is not None for cell in row)
            ]
            assert xlsx_rows == [
                ["2026-03-05", "expense", 1, "Food"],
                ["2026-03-01", "transfer", None, "Transfer"],
                ["2026-03-10", "income", 2, "Salary"],
            ]
        finally:
            wb.close()
    finally:
        for p in (csv_path, xlsx_path):
            if p.exists():
                for _ in range(5):
                    try:
                        os.unlink(p)
                        break
                    except PermissionError:
                        time.sleep(0.1)


def test_export_full_backup_preserves_storage_metadata() -> None:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".json") as json_tmp:
        json_path = Path(json_tmp.name)

    try:
        exporters.export_full_backup(
            str(json_path),
            wallets=[],
            records=[],
            mandatory_expenses=[],
            distribution_items=[],
            distribution_subitems=[],
            distribution_snapshots=[],
            transfers=[],
            readonly=True,
            storage_mode="sqlite",
        )

        import json

        payload = json.loads(json_path.read_text(encoding="utf-8"))
        assert payload["meta"]["storage"] == "sqlite"
    finally:
        if json_path.exists():
            os.unlink(json_path)


def test_import_full_backup_exposes_debts_and_payments_without_breaking_legacy_unpack() -> None:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".json") as json_tmp:
        json_path = Path(json_tmp.name)

    try:
        debt = Debt(
            id=1,
            contact_name="Alice",
            kind=DebtKind.LOAN,
            total_amount_minor=100000,
            remaining_amount_minor=40000,
            currency="KZT",
            interest_rate=0.0,
            status=DebtStatus.OPEN,
            created_at="2026-04-01",
        )
        payment = DebtPayment(
            id=1,
            debt_id=1,
            record_id=1,
            operation_type=DebtOperationType.LOAN_COLLECT,
            principal_paid_minor=60000,
            is_write_off=False,
            payment_date="2026-04-02",
        )
        sample_record = IncomeRecord(
            id=1,
            date="2026-04-02",
            wallet_id=1,
            related_debt_id=1,
            amount_original=600.0,
            currency="KZT",
            rate_at_operation=1.0,
            amount_base=600.0,
            category="Loan payment",
        )

        exporters.export_full_backup(
            str(json_path),
            wallets=[],
            records=[sample_record],
            mandatory_expenses=[],
            debts=[debt],
            debt_payments=[payment],
            distribution_items=[],
            distribution_subitems=[],
            distribution_snapshots=[],
            transfers=[],
            readonly=False,
            storage_mode="sqlite",
        )

        imported = importers.import_full_backup(str(json_path), force=True)
        wallets, records, mandatory, transfers, summary = imported

        assert len(wallets) == 1
        assert isinstance(wallets[0], Wallet)
        assert wallets[0].system is True
        assert len(records) == 1
        assert mandatory == []
        assert transfers == []
        assert summary[1] == 0
        assert len(imported.debts) == 1
        assert len(imported.debt_payments) == 1
        assert imported.debts[0].contact_name == "Alice"
        assert imported.debt_payments[0].record_id == 1
    finally:
        if json_path.exists():
            os.unlink(json_path)


def test_export_full_backup_includes_debts_and_debt_payments() -> None:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".json") as json_tmp:
        json_path = Path(json_tmp.name)

    try:
        exporters.export_full_backup(
            str(json_path),
            wallets=[],
            records=[],
            mandatory_expenses=[],
            debts=[
                Debt(
                    id=1,
                    contact_name="Alice",
                    kind=DebtKind.DEBT,
                    total_amount_minor=50000,
                    remaining_amount_minor=25000,
                    currency="KZT",
                    interest_rate=0.0,
                    status=DebtStatus.OPEN,
                    created_at="2026-03-01",
                )
            ],
            debt_payments=[
                DebtPayment(
                    id=1,
                    debt_id=1,
                    record_id=None,
                    operation_type=DebtOperationType.DEBT_FORGIVE,
                    principal_paid_minor=25000,
                    is_write_off=True,
                    payment_date="2026-03-10",
                )
            ],
            distribution_items=[],
            distribution_subitems=[],
            distribution_snapshots=[],
            transfers=[],
            readonly=True,
            storage_mode="sqlite",
        )

        payload = json.loads(json_path.read_text(encoding="utf-8"))
        data = payload["data"]
        assert data["debts"][0]["contact_name"] == "Alice"
        assert data["debt_payments"][0]["operation_type"] == "debt_forgive"
        assert data["debt_payments"][0]["is_write_off"] is True
    finally:
        if json_path.exists():
            os.unlink(json_path)
